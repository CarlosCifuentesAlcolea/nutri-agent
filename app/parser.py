from openpyxl import load_workbook
import re


def numero_a_float(valor):
    if valor is None:
        return None

    texto = str(valor).strip().replace(",", ".")

    if "/" in texto:
        numerador, denominador = texto.split("/", 1)
        return float(numerador) / float(denominador)

    return float(texto)


def _parsear_alimento(texto):
    texto = str(texto).strip()
    patron = r"^(?P<cantidad>\d+(?:[,.]\d+)?(?:/\d+)?)\s*(?P<unidad>g|gr|ml)?\s+(?P<alimento>.+)$"
    coincidencia = re.match(patron, texto, re.IGNORECASE)

    if not coincidencia:
        return {
            "cantidad": None,
            "unidad": None,
            "alimento": texto,
            "descripcion": texto
        }

    cantidad = coincidencia.group("cantidad").replace(",", ".")

    return {
        "cantidad": cantidad,
        "unidad": coincidencia.group("unidad") or "unidad",
        "alimento": coincidencia.group("alimento").strip(),
        "descripcion": texto
    }


def extraer_objetivos(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)

    ws = wb["Dieta"]

    hc_texto = str(ws["C6"].value)
    proteina = float(ws["D6"].value)
    grasa_texto = str(ws["E6"].value)

    hc_numeros = [int(x) for x in re.findall(r"\d+", hc_texto)]
    grasa_numeros = [int(x) for x in re.findall(r"\d+", grasa_texto)]

    objetivos = {
        "entrenamiento": {
            "hc": hc_numeros[0],
            "proteina": proteina,
            "grasa": grasa_numeros[0]
        },
        "descanso": {
            "hc": hc_numeros[1],
            "proteina": proteina,
            "grasa": grasa_numeros[1]
        }
    }

    return objetivos

def buscar_equivalencias(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)

    ws = wb["Dieta"]

    for fila in range(1, 120):
        for col in range(1, 10):

            valor = ws.cell(fila, col).value

            if valor:
                texto = str(valor)

                if "1 HIDRATOS" in texto:
                    print("HC encontrado:", fila, col)

                if "1 PROTEÍNA" in texto:
                    print("PROTEINA encontrada:", fila, col)

                if "1 GRASA" in texto:
                    print("GRASA encontrada:", fila, col)

def mostrar_equivalencias(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)

    ws = wb["Dieta"]

    for fila in range(20, 60):
        valores = []

        for col in range(2, 6):
            valor = ws.cell(fila, col).value

            if valor is not None:
                valores.append(str(valor))

        if valores:
            print(f"FILA {fila}: {valores}")


def _es_encabezado_equivalencia(texto):
    texto = texto.upper()
    palabras_encabezado = ("HIDRATOS", "PROTE", "GRASA", "CARBOHIDRATOS")
    return any(palabra in texto for palabra in palabras_encabezado)


def _parsear_macros_intercambio(intercambio):
    macros = {"hc": 0.0, "proteina": 0.0, "grasa": 0.0}
    texto = intercambio.lower()
    patron = r"(?:(\d+(?:[,.]\d+)?(?:/\d+)?)\s*)?(hidratos?|carbohidratos?|proteinas?|grasa|grasas)"

    for cantidad, macro in re.findall(patron, texto):
        valor = numero_a_float(cantidad) if cantidad else 1.0

        if macro.startswith(("hidrato", "carbohidrato")):
            macros["hc"] += valor
        elif macro.startswith("proteina"):
            macros["proteina"] += valor
        elif macro.startswith("grasa"):
            macros["grasa"] += valor

    return macros


def _cargar_equivalencias_columna(ruta_excel, columna, secciones):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb["Dieta"]

    equivalencias = []

    for inicio, fin, intercambio in secciones:
        for fila in range(inicio, fin + 1):
            valor = ws.cell(fila, columna).value

            if valor is None:
                continue

            texto = str(valor).strip()

            if _es_encabezado_equivalencia(texto):
                continue

            alimento = _parsear_alimento(texto)
            alimento["fila_excel"] = fila
            alimento["intercambio"] = intercambio
            alimento["macros"] = _parsear_macros_intercambio(intercambio)

            equivalencias.append(alimento)

    return equivalencias


def cargar_hidratos(ruta_excel):
    return _cargar_equivalencias_columna(
        ruta_excel,
        columna=3,
        secciones=[
            (25, 48, "1 hidrato"),
            (51, 62, "1/2 proteina + 1/2 hidrato"),
        ],
    )


def cargar_proteinas(ruta_excel):
    return _cargar_equivalencias_columna(
        ruta_excel,
        columna=4,
        secciones=[
            (25, 48, "1 proteina"),
            (51, 73, "1/2 proteina + 1/2 grasa"),
        ],
    )


def cargar_grasas(ruta_excel):
    return _cargar_equivalencias_columna(
        ruta_excel,
        columna=5,
        secciones=[
            (25, 48, "1 grasa"),
            (51, 65, "0.25 hidratos + 0.25 proteina + 0.5 grasa"),
            (67, 68, "1/2 hidratos + 1/2 grasa"),
        ],
    )


def cargar_equivalencias(ruta_excel):
    return {
        "hidratos": cargar_hidratos(ruta_excel),
        "proteinas": cargar_proteinas(ruta_excel),
        "grasas": cargar_grasas(ruta_excel),
    }
